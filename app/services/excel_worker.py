"""app/services/excel_worker.py — фоновая запись в Excel после коммита.

FIX: stop() теперь дожидается флашинга очереди перед завершением потока,
     чтобы незаписанные строки не терялись при штатной остановке бота.
"""
from __future__ import annotations

import logging
import queue
import threading
from dataclasses import dataclass
from pathlib import Path

from app.core.settings import settings
from app.core.money import format_eur

logger = logging.getLogger("salon.excel")

_queue: queue.Queue = queue.Queue(maxsize=500)
_STOP = object()
_thread: threading.Thread | None = None


@dataclass
class ExcelRow:
    kind: str; name: str; phone: str; service: str
    date: str; time: str; price: int; status: str


def enqueue(row: ExcelRow):
    try:
        _queue.put_nowait(row)
    except queue.Full:
        logger.warning("Excel queue full, row dropped: %s %s", row.date, row.name)


def _ensure(path: str):
    if not Path(path).exists():
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Записи"
            ws.append(["Тип", "Имя", "Телефон", "Услуга", "Дата", "Время", "Цена (EUR)", "Статус"])
            wb.save(path)
        except Exception as e:
            logger.error("Excel create: %s", e)


def _write(path: str, row: ExcelRow):
    try:
        from filelock import FileLock
        from openpyxl import load_workbook
        with FileLock(path + ".lock", timeout=10):
            wb = load_workbook(path)
            wb.active.append([
                row.kind, row.name, row.phone, row.service,
                row.date, row.time, format_eur(row.price), row.status,
            ])
            wb.save(path)
    except Exception as e:
        logger.error("Excel write: %s", e)


def _worker():
    path = settings.EXCEL_PATH
    _ensure(path)
    logger.info("Excel worker started")
    while True:
        try:
            item = _queue.get(timeout=2)
            if item is _STOP:
                # FIX: дочищаем остаток очереди перед выходом
                while True:
                    try:
                        leftover = _queue.get_nowait()
                        if leftover is not _STOP:
                            _write(path, leftover)
                    except queue.Empty:
                        break
                logger.info("Excel worker stopped (queue flushed)")
                break
            _write(path, item)
        except queue.Empty:
            continue
        except Exception as e:
            logger.error("Excel worker: %s", e)


def start() -> threading.Thread:
    global _thread
    t = threading.Thread(target=_worker, daemon=True, name="excel-worker")
    t.start()
    _thread = t
    return t


def stop(timeout: float = 10.0):
    """Отправляет сигнал остановки и ждёт завершения потока."""
    global _thread
    _queue.put(_STOP)
    if _thread and _thread.is_alive():
        _thread.join(timeout=timeout)
    _thread = None
